"""Generate ShelvesFinder manual test Excel workbook."""

from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    raise SystemExit("Run: pip install openpyxl")

OUTPUT = Path(__file__).resolve().parent.parent / "ShelvesFinder_Test_Cases.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="CCCCCC")


def style_header(ws, row: int, ncol: int) -> None:
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_widths(ws, widths: dict) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def add_sheet_instructions(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "How to Use"
    rows = [
        ("ShelvesFinder — Manual Test Workbook", ""),
        ("", ""),
        ("1. Go to sheet 'Test Cases'", "Main sheet — run tests and fill yellow columns."),
        ("2. Sheet 'Example URLs'", "Copy/paste Walmart product links."),
        ("3. Sheet 'Env Checklist'", "Confirm API keys before live UI tests."),
        ("4. Sheet 'Automated Tests'", "Run pytest — no manual Pass/Fail needed."),
        ("", ""),
        ("How to start the app (UI tests)", ""),
        ("Step 1", "Open terminal in folder: backend"),
        ("Step 2", "uvicorn app.main:app --reload"),
        ("Step 3", "Open frontend in browser (or http://localhost:8000)"),
        ("", ""),
        ("Run automated tests", ""),
        ("Command", "cd backend"),
        ("", "pip install pytest pytest-asyncio"),
        ("", "python -m pytest tests/ -v"),
        ("", ""),
        ("Pass/Fail column values", "Pass | Fail | Blocked | Not Run"),
        ("Blocked", "Use when API key missing, server down, or Walmart bot block prevents fair test."),
    ]
    ws["A1"].font = TITLE_FONT
    for i, (a, b) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=b)
        ws.cell(row=i, column=1).alignment = WRAP
        ws.cell(row=i, column=2).alignment = WRAP
    set_widths(ws, {"A": 28, "B": 70})


def test_cases_data():
    broccoli = (
        "https://www.walmart.com/ip/Simplot-IQF-Broccoli-Florets-32-oz-package-"
        "12-packages-per-case/504628592?classType=REGULAR&from=/search"
    )
    sourdough = (
        "https://www.walmart.com/ip/Essential-Hawaiian-Sliced-Sourdough-Loaf-"
        "Non-GMO-16-oz/12928764204?classType=REGULAR&adsRedirect=true"
    )
    dress = (
        "https://www.walmart.com/ip/Htigea-Wedding-Guest-Midi-Dress-for-Women-"
        "Long-Sleeve-V-Neck-Tie-Waist-Bodycon-Dresses-Elegant-Formal-Party-"
        "Cocktail-Dress-Black-XL/19307773882"
    )
    whole_milk_browse = (
        "https://www.walmart.com/browse/food/whole-milk/"
        "976759_9176907_4405816_6541475_6877499"
    )

    return [
        {
            "id": "TC-URL-01",
            "category": "URL Reference",
            "name": "Htigea dress — product URL parsing",
            "priority": "Medium",
            "where": "pytest (automated) OR scrape API",
            "mode": "N/A",
            "auto": "N/A",
            "input": dress,
            "settings": "Expected product ID: 19307773882",
            "preconditions": "None for URL parse; API needs server running",
            "steps": "1. Run pytest test_product_urls OR\n2. POST /analyze/step/scrape with URL",
            "expected": "ID = 19307773882; slug title contains dress keywords",
        },
        {
            "id": "TC-URL-02",
            "category": "URL Reference",
            "name": "Sourdough — URL with query params",
            "priority": "Medium",
            "where": "Browser UI or API",
            "mode": "Basic",
            "auto": "ON",
            "input": sourdough,
            "settings": "Expected product ID: 12928764204",
            "preconditions": "Server + LLM key",
            "steps": "1. Paste URL in app\n2. Analyze (Basic Auto)\n3. Check scrape step output",
            "expected": "ID = 12928764204 despite ?classType=REGULAR&adsRedirect=true",
        },
        {
            "id": "TC-URL-03",
            "category": "URL Reference",
            "name": "Simplot broccoli — case pack product",
            "priority": "High",
            "where": "Browser UI",
            "mode": "Basic or Advance",
            "auto": "Either",
            "input": broccoli,
            "settings": "Expected product ID: 504628592",
            "preconditions": "Server + SERPER + LLM keys",
            "steps": "1. Paste URL\n2. Run analysis\n3. Verify ID in scrape output",
            "expected": "ID = 504628592; keywords are shelf-style (frozen broccoli), not full case title",
        },
        {
            "id": "TC-URL-04",
            "category": "URL Reference",
            "name": "Whole milk — browse shelf URL (reference only)",
            "priority": "Low",
            "where": "N/A — do not paste as input",
            "mode": "N/A",
            "auto": "N/A",
            "input": whole_milk_browse,
            "settings": "This is a /browse/ shelf URL, not /ip/ product",
            "preconditions": "N/A",
            "steps": "Use only to understand Search output format.\nInput must be /ip/ product URLs.",
            "expected": "App discovers similar /browse/ URLs during Search step",
        },
        {
            "id": "TC-V1-AUTO-01",
            "category": "UI — Basic",
            "name": "Full auto pipeline — broccoli",
            "priority": "High",
            "where": "Browser UI — http://localhost:8000",
            "mode": "Basic",
            "auto": "ON",
            "input": broccoli,
            "settings": "Default settings",
            "preconditions": "Server running; SERPER_API_KEY + LLM key in backend/.env",
            "steps": "1. Select Basic mode\n2. Auto toggle ON\n3. Paste broccoli URL\n4. Click Analyze\n5. Watch: Scrape → Keywords → Search → Evaluation → Visibility",
            "expected": "All 5 steps complete.\nScrape: title has Broccoli or slug; ID 504628592.\nKeywords: 6-8 unbranded shelf terms.\nSearch: browse URLs if Serper works.\nEvaluation: ranked list + confidence.\nVisibility: found/missing stats.",
        },
        {
            "id": "TC-V1-AUTO-02",
            "category": "UI — Basic",
            "name": "Sourdough with query string",
            "priority": "Medium",
            "where": "Browser UI",
            "mode": "Basic",
            "auto": "ON",
            "input": sourdough,
            "settings": "Default",
            "preconditions": "Server + keys",
            "steps": "1. Basic + Auto ON\n2. Paste sourdough URL\n3. Analyze\n4. Open scrape output",
            "expected": "Product ID = 12928764204 in scrape and shelf check",
        },
        {
            "id": "TC-V1-MANUAL-01",
            "category": "UI — Basic",
            "name": "Manual step-by-step — edit keywords",
            "priority": "High",
            "where": "Browser UI",
            "mode": "Basic",
            "auto": "OFF",
            "input": dress,
            "settings": "Edit keywords on step 2",
            "preconditions": "Server + LLM key",
            "steps": "1. Basic, Auto OFF\n2. Paste dress URL → Analyze\n3. Scrape → Continue\n4. Edit unbranded keywords to:\n   cocktail dresses\n   party dresses\n5. Continue through Search → Evaluate → Visibility",
            "expected": "Search uses YOUR keywords.\nEvaluate ranks dress-related browse pages.",
        },
        {
            "id": "TC-V1-NEG-01",
            "category": "UI — Basic",
            "name": "Invalid URL validation",
            "priority": "Medium",
            "where": "Browser UI",
            "mode": "Basic",
            "auto": "Either",
            "input": "https://www.amazon.com/dp/B000 (or leave empty)",
            "settings": "None",
            "preconditions": "Server running",
            "steps": "1. Paste Amazon URL or empty field\n2. Click Analyze",
            "expected": "Error message shown; analysis does not start",
        },
        {
            "id": "TC-V2-01",
            "category": "UI — Advance",
            "name": "Default agent run — broccoli",
            "priority": "High",
            "where": "Browser UI",
            "mode": "Advance",
            "auto": "N/A",
            "input": broccoli,
            "settings": "Max rounds=5, Target missing=3, Budget=$0.50",
            "preconditions": "Server + SERPER + LLM keys",
            "steps": "1. Select Advance\n2. Paste broccoli URL\n3. Open Settings — confirm defaults\n4. Analyze\n5. Watch reasoning log until complete",
            "expected": "Setup: scrape + keywords.\nLoop shows search, evaluate, check_shelf.\nStops at 3 missing OR limits.\nCost near or under budget.",
        },
        {
            "id": "TC-V2-02",
            "category": "UI — Advance",
            "name": "Agent context instructions",
            "priority": "Medium",
            "where": "Browser UI",
            "mode": "Advance",
            "auto": "N/A",
            "input": broccoli,
            "settings": "Agent context: Focus on frozen vegetables and frozen broccoli only. Skip snacks and dairy.",
            "preconditions": "Server + keys",
            "steps": "1. Advance mode\n2. Paste URL\n3. Enter agent context in text box\n4. Analyze\n5. Read reasoning log keywords/searches",
            "expected": "Keywords and searches lean toward frozen produce (check log subjectively)",
        },
        {
            "id": "TC-V2-03",
            "category": "UI — Advance",
            "name": "Branded results enabled",
            "priority": "Medium",
            "where": "Browser UI",
            "mode": "Advance",
            "auto": "N/A",
            "input": broccoli,
            "settings": "Include Branded Results = ON",
            "preconditions": "Server + keys",
            "steps": "1. Advance → Settings\n2. Check Include Branded Results\n3. Analyze broccoli URL\n4. Review keyword list in log/setup",
            "expected": "Keywords include brand terms (e.g. Simplot + product type) plus unbranded shelf terms; brand-specific category shelves may appear in results",
        },
        {
            "id": "TC-V2-04",
            "category": "UI — Advance",
            "name": "Low budget early stop",
            "priority": "Low",
            "where": "Browser UI",
            "mode": "Advance",
            "auto": "N/A",
            "input": broccoli,
            "settings": "Budget=$0.01, Max rounds=20",
            "preconditions": "Server + keys",
            "steps": "1. Advance → Settings\n2. Set Budget 0.01, Max rounds 20\n3. Analyze\n4. Note stop reason in log",
            "expected": "Run stops early; log mentions budget limit",
        },
        {
            "id": "TC-API-01",
            "category": "API",
            "name": "Scrape endpoint only",
            "priority": "High",
            "where": "Postman / curl — http://localhost:8000",
            "mode": "N/A",
            "auto": "N/A",
            "input": "POST /analyze/step/scrape\nBody: {\"url\": \"<broccoli URL>\"}",
            "settings": "Content-Type: application/json",
            "preconditions": "uvicorn app.main:app --reload",
            "steps": "curl -X POST http://localhost:8000/analyze/step/scrape -H \"Content-Type: application/json\" -d \"{\\\"url\\\": \\\"<broccoli URL>\\\"}\"",
            "expected": "JSON with title, id=504628592, brand (may be empty if bot blocked)",
        },
        {
            "id": "TC-API-02",
            "category": "API",
            "name": "v1 full stream (SSE)",
            "priority": "Medium",
            "where": "curl / Postman",
            "mode": "Basic",
            "auto": "N/A",
            "input": "GET /analyze/stream?url=<sourdough URL>",
            "settings": "SSE stream",
            "preconditions": "Server + keys",
            "steps": "curl -N \"http://localhost:8000/analyze/stream?url=<encoded sourdough URL>\"",
            "expected": "Events: scraping, keywords, search, evaluation, visibility, done",
        },
        {
            "id": "TC-API-03",
            "category": "API",
            "name": "v2 health check",
            "priority": "Low",
            "where": "curl / browser",
            "mode": "N/A",
            "auto": "N/A",
            "input": "GET http://localhost:8000/v2/health",
            "settings": "None",
            "preconditions": "Server running",
            "steps": "curl http://localhost:8000/v2/health",
            "expected": "HTTP 200 OK",
        },
        {
            "id": "TC-AUTO-01",
            "category": "Automated (pytest)",
            "name": "All unit tests",
            "priority": "High",
            "where": "Terminal — backend folder",
            "mode": "N/A",
            "auto": "N/A",
            "input": "python -m pytest tests/ -v",
            "settings": "pip install pytest pytest-asyncio",
            "preconditions": "Python env with app dependencies",
            "steps": "cd backend\npip install pytest pytest-asyncio\npython -m pytest tests/ -v",
            "expected": "All tests pass (50+)",
        },
    ]


def add_sheet_test_cases(wb: Workbook) -> None:
    ws = wb.create_sheet("Test Cases")
    headers = [
        "Test ID",
        "Category",
        "Test Name",
        "Priority",
        "Where to Run",
        "Mode",
        "Auto On/Off",
        "Input URL / Command",
        "Settings / Config",
        "Preconditions",
        "Test Steps",
        "Expected Result",
        "Actual Result (you fill)",
        "Pass / Fail",
        "Date Tested",
        "Tester Name",
        "Notes",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for tc in test_cases_data():
        ws.append([
            tc["id"],
            tc["category"],
            tc["name"],
            tc["priority"],
            tc["where"],
            tc["mode"],
            tc["auto"],
            tc["input"],
            tc["settings"],
            tc["preconditions"],
            tc["steps"],
            tc["expected"],
            "",
            "Not Run",
            "",
            "",
            "",
        ])

    fill_yellow = PatternFill("solid", fgColor="FFF9E5")
    result_cols = [13, 14, 15, 16, 17]  # Actual, Pass/Fail, Date, Tester, Notes
    for row in range(2, ws.max_row + 1):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = WRAP
            if col in result_cols:
                cell.fill = fill_yellow

    dv = DataValidation(
        type="list",
        formula1='"Pass,Fail,Blocked,Not Run"',
        allow_blank=True,
    )
    dv.error = "Choose: Pass, Fail, Blocked, or Not Run"
    ws.add_data_validation(dv)
    dv.add(f"N2:N{ws.max_row}")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    set_widths(ws, {
        "A": 14, "B": 14, "C": 28, "D": 10, "E": 22, "F": 10, "G": 10,
        "H": 40, "I": 28, "J": 28, "K": 36, "L": 36, "M": 28, "N": 12,
        "O": 14, "P": 14, "Q": 24,
    })


def add_sheet_urls(wb: Workbook) -> None:
    ws = wb.create_sheet("Example URLs")
    headers = ["Ref ID", "Product", "Full URL", "Expected Product ID", "URL Type"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    rows = [
        ("TC-URL-01", "Htigea dress",
         "https://www.walmart.com/ip/Htigea-Wedding-Guest-Midi-Dress-for-Women-Long-Sleeve-V-Neck-Tie-Waist-Bodycon-Dresses-Elegant-Formal-Party-Cocktail-Dress-Black-XL/19307773882",
         "19307773882", "Product /ip/ — use as INPUT"),
        ("TC-URL-02", "Essential sourdough",
         "https://www.walmart.com/ip/Essential-Hawaiian-Sliced-Sourdough-Loaf-Non-GMO-16-oz/12928764204?classType=REGULAR&adsRedirect=true",
         "12928764204", "Product /ip/ — use as INPUT"),
        ("TC-URL-03", "Simplot broccoli",
         "https://www.walmart.com/ip/Simplot-IQF-Broccoli-Florets-32-oz-package-12-packages-per-case/504628592?classType=REGULAR&from=/search",
         "504628592", "Product /ip/ — use as INPUT"),
        ("TC-URL-04", "Whole milk shelf",
         "https://www.walmart.com/browse/food/whole-milk/976759_9176907_4405816_6541475_6877499",
         "(category IDs — not product)", "Browse /browse/ — OUTPUT from Search only"),
    ]
    for r in rows:
        ws.append(list(r))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = WRAP
    ws.freeze_panes = "A2"
    set_widths(ws, {"A": 12, "B": 18, "C": 80, "D": 22, "E": 32})


def add_sheet_env(wb: Workbook) -> None:
    ws = wb.create_sheet("Env Checklist")
    headers = [
        "Variable", "Required For", "Configured? (Y/N)",
        "Value Present? (Y/N)", "Notes",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    env_rows = [
        ("OPENAI_API_KEY", "Keywords, embeddings, optional orchestrator", "", "", ""),
        ("ANTHROPIC_API_KEY", "Keywords + Advance orchestrator (if using Claude)", "", "", ""),
        ("SERPER_API_KEY", "Search step — find browse URLs", "", "", ""),
        ("WEBSCRAPING_API_KEY", "Optional — better Walmart scrape/shelf fetch", "", "", ""),
        ("LLM_PROVIDER", "openai or claude in .env", "", "", ""),
    ]
    for r in env_rows:
        ws.append(list(r))
    dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("C2:D6")
    set_widths(ws, {"A": 24, "B": 40, "C": 16, "D": 18, "E": 30})


def add_sheet_automated(wb: Workbook) -> None:
    ws = wb.create_sheet("Automated Tests")
    headers = [
        "Test File", "What It Tests", "Command", "Last Run Date",
        "Pass/Fail", "Notes",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    rows = [
        ("test_product_urls.py", "ID + slug for dress, sourdough, broccoli",
         "python -m pytest tests/test_product_urls.py -v", "", "", ""),
        ("test_search_api.py", "Browse URL filter, Serper mock",
         "python -m pytest tests/test_search_api.py -v", "", "", ""),
        ("test_search_agent.py", "URL dedup, host guard",
         "python -m pytest tests/test_search_agent.py -v", "", "", ""),
        ("test_similarity.py", "Cosine + keyword overlap fallback",
         "python -m pytest tests/test_similarity.py -v", "", "", ""),
        ("test_shelf_checker.py", "Found / missing / 404 pages",
         "python -m pytest tests/test_shelf_checker.py -v", "", "", ""),
        ("test_evaluation_agent.py", "Browse page ranking",
         "python -m pytest tests/test_evaluation_agent.py -v", "", "", ""),
        ("test_pipeline_htigea.py", "Mocked end-to-end pipeline",
         "python -m pytest tests/test_pipeline_htigea.py -v", "", "", ""),
        ("ALL", "Full suite (~50 tests)",
         "python -m pytest tests/ -v", "", "", ""),
    ]
    for r in rows:
        ws.append(list(r))
    dv = DataValidation(type="list", formula1='"Pass,Fail,Not Run"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("E2:E9")
    set_widths(ws, {"A": 28, "B": 36, "C": 48, "D": 14, "E": 12, "F": 24})


def main() -> None:
    wb = Workbook()
    add_sheet_instructions(wb)
    add_sheet_test_cases(wb)
    add_sheet_urls(wb)
    add_sheet_env(wb)
    add_sheet_automated(wb)
    wb.save(OUTPUT)
    print(f"Created: {OUTPUT}")


if __name__ == "__main__":
    main()
